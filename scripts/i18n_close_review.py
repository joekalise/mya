#!/usr/bin/env python3
"""Close out reviewed translations: import into es.json, drop them from the
pending queue, and advance the review baseline so they aren't re-flagged.

Usage:
  python3 scripts/i18n_close_review.py [reviewed.xlsx]

Defaults to ~/Downloads/mya_es_review.xlsx. A row is "closed" only if its
Reviewed (Y/N) column is marked Y and its Target (es) cell is non-empty.
Target is pre-filled from the current es.json for context, so a non-empty
Target alone doesn't mean a human looked at it, the Reviewed mark does.
Rows not marked stay in the pending queue and in the regenerated .xlsx.
"""
import subprocess
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

from i18n_lib import (
    BASELINE_PATH,
    DEFAULT_XLSX,
    EN_PATH,
    ES_PATH,
    get_nested,
    load_json,
    read_pending,
    set_nested,
    write_json,
    write_pending,
)


def main():
    if len(sys.argv) not in (1, 2):
        print("Usage: i18n_close_review.py [reviewed.xlsx]")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1]) if len(sys.argv) == 2 else DEFAULT_XLSX
    es = load_json(ES_PATH)
    en = load_json(EN_PATH)
    baseline = load_json(BASELINE_PATH) if BASELINE_PATH.exists() else {}

    wb = load_workbook(xlsx_path)
    ws = wb.active

    closed = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        ns, key, _source, target, reviewed = row[0], row[1], row[2], row[3], row[4]
        if str(reviewed or "").strip().lower() not in ("y", "yes"):
            continue
        if target is None or str(target).strip() == "":
            continue
        full_key = f"{ns}.{key}"
        set_nested(es, full_key, str(target))
        set_nested(baseline, full_key, get_nested(en, full_key))
        closed.append(full_key)

    if not closed:
        print("No rows marked Reviewed with a Target (es) value, nothing to close.")
        return

    write_json(ES_PATH, es)
    write_json(BASELINE_PATH, baseline)

    pending = read_pending()
    closed_set = set(closed)
    remaining = [k for k in pending if k not in closed_set]
    write_pending(remaining)

    print(f"Closed {len(closed)} key(s):")
    for k in closed:
        print(f"  {k}")
    print(f"{len(remaining)} key(s) still pending.")

    subprocess.run([sys.executable, str(Path(__file__).resolve().parent / "i18n_xlsx_export.py")], check=True)


if __name__ == "__main__":
    main()

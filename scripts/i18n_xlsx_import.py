#!/usr/bin/env python3
"""Import reviewed translations from an .xlsx (produced by i18n_xlsx_export.py)
back into es.json.

Usage:
  python3 scripts/i18n_xlsx_import.py [reviewed.xlsx]

Defaults to ~/Downloads/mya_es_review.xlsx. Expects columns: Namespace, Key,
Source (en-GB), Target (es), matching the export format. Only updates a key
if the Target cell is non-empty and differs from the current value, and
reports every change it makes.

This only writes es.json. To also clear reviewed keys out of the pending
queue and update the review baseline, use i18n_close_review.py instead.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

from i18n_lib import DEFAULT_XLSX, ES_PATH, get_nested, load_json, set_nested, write_json


def main():
    if len(sys.argv) not in (1, 2):
        print("Usage: i18n_xlsx_import.py [reviewed.xlsx]")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1]) if len(sys.argv) == 2 else DEFAULT_XLSX
    es = load_json(ES_PATH)

    wb = load_workbook(xlsx_path)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    changed = []
    unchanged = 0

    for row in rows:
        if not row or row[0] is None:
            continue
        ns, key, _source, target = row[0], row[1], row[2], row[3]
        if target is None or str(target).strip() == "":
            continue
        target = str(target)
        full_key = f"{ns}.{key}"
        current = get_nested(es, full_key)
        if current == target:
            unchanged += 1
            continue
        old = current
        set_nested(es, full_key, target)
        changed.append((full_key, old, target))

    if changed:
        write_json(ES_PATH, es)

    print(f"Updated {len(changed)} key(s), {unchanged} already matched.")
    for k, old, new in changed:
        print(f"  {k}:")
        print(f"    - {old!r}")
        print(f"    + {new!r}")


if __name__ == "__main__":
    main()

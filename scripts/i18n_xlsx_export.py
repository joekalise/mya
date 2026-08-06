#!/usr/bin/env python3
"""Export the pending review queue (source EN + target ES) to an .xlsx.

Usage:
  python3 scripts/i18n_xlsx_export.py [key_list.txt] [output.xlsx]

Defaults to scripts/i18n_pending_review.txt -> ~/Downloads/mya_es_review.xlsx.
The key list file has one dotted "namespace.key" path per line (blank lines
ignored), dotted paths reach into nested namespaces like "onboarding.age_range.title".
Pair this with i18n_xlsx_import.py to bring reviewed/edited translations back
into es.json afterwards.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from i18n_lib import DEFAULT_XLSX, EN_PATH, ES_PATH, PENDING_PATH, get_nested, load_json


def main():
    if len(sys.argv) not in (1, 3):
        print("Usage: i18n_xlsx_export.py [key_list.txt] [output.xlsx]")
        sys.exit(1)

    key_list_path = Path(sys.argv[1]) if len(sys.argv) == 3 else PENDING_PATH
    out_path = Path(sys.argv[2]) if len(sys.argv) == 3 else DEFAULT_XLSX

    keys = [line.strip() for line in key_list_path.read_text().splitlines() if line.strip()]

    en = load_json(EN_PATH)
    es = load_json(ES_PATH)

    wb = Workbook()
    ws = wb.active
    ws.title = "es review"

    headers = ["Namespace", "Key", "Source (en-GB)", "Target (es)", "Reviewed (Y/N)"]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="F97316")

    missing = []
    for full_key in keys:
        ns, key = full_key.split(".", 1)
        src = get_nested(en, full_key)
        tgt = get_nested(es, full_key)
        if src is None:
            missing.append(full_key)
            continue
        # Target is pre-filled with whatever's already in es.json for context,
        # but that's not a review, only a checked "Reviewed" column closes a row.
        ws.append([ns, key, src, tgt if tgt is not None else "", ""])

    widths = [16, 26, 55, 55, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row[2:4]:
            cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(out_path)

    print(f"Wrote {len(keys) - len(missing)} rows to {out_path}")
    if missing:
        print(f"Skipped {len(missing)} keys not found in en-GB.json: {missing}")


if __name__ == "__main__":
    main()
